#!/usr/bin/env python3
"""COCHING DB → Excel 내보내기 (깔끔한 포맷)"""
import json, os
import psycopg2
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from datetime import datetime

# DB 연결
conn = psycopg2.connect(
    host="127.0.0.1", port=5432,
    dbname="coching_db", user="coching_user",
    password=os.environ.get("COCHING_DB_PASS", "changeme")
)

wb = Workbook()

# ── 스타일 정의 ──────────────────────────────────────────────────────────────
HEADER_FONT = Font(name="맑은 고딕", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SUB_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
CELL_FONT = Font(name="맑은 고딕", size=10)
TITLE_FONT = Font(name="맑은 고딕", bold=True, size=14, color="2F5496")
SUBTITLE_FONT = Font(name="맑은 고딕", size=10, color="808080")
THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)

def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

def style_data(ws, row, cols, alt=False):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = CELL_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        if alt:
            cell.fill = PatternFill(start_color="F2F7FC", end_color="F2F7FC", fill_type="solid")

def add_title(ws, title, subtitle=""):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = TITLE_FONT
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 35
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
        cell2 = ws.cell(row=2, column=1, value=subtitle)
        cell2.font = SUBTITLE_FONT
        ws.row_dimensions[2].height = 20

def auto_width(ws, cols, max_width=50):
    for c in range(1, cols + 1):
        max_len = 0
        col_letter = get_column_letter(c)
        for row in ws.iter_rows(min_col=c, max_col=c, values_only=False):
            for cell in row:
                if cell.value:
                    lines = str(cell.value).split('\n')
                    for line in lines:
                        # 한글은 약 2배 폭
                        length = sum(2 if ord(ch) > 127 else 1 for ch in line)
                        max_len = max(max_len, length)
        ws.column_dimensions[col_letter].width = min(max_len + 3, max_width)


# ════════════════════════════════════════════════════════════════════════════
# Sheet 1: 원료 마스터
# ════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "원료 마스터"
ws1.sheet_properties.tabColor = "2F5496"

cur = conn.cursor()
cur.execute("""
    SELECT id, inci_name, korean_name, cas_number, ingredient_type,
           origin, description, source
    FROM ingredient_master
    ORDER BY ingredient_type, inci_name
""")
rows = cur.fetchall()

add_title(ws1, "화장품 원료 마스터 DB", f"총 {len(rows)}건 | 추출일: {datetime.now().strftime('%Y-%m-%d %H:%M')}")

headers = ["No.", "INCI명 (국제명칭)", "한국어명", "CAS 번호", "원료 유형", "원산지", "설명", "데이터 출처"]
START_ROW = 4
for c, h in enumerate(headers, 1):
    ws1.cell(row=START_ROW, column=c, value=h)
style_header(ws1, START_ROW, len(headers))
ws1.row_dimensions[START_ROW].height = 30

# 유형 한글 매핑
TYPE_KR = {
    "EMULSIFIER": "유화제", "SURFACTANT": "계면활성제", "HUMECTANT": "보습제",
    "EMOLLIENT": "연화제", "PRESERVATIVE": "방부제", "ANTIOXIDANT": "항산화제",
    "UV_FILTER": "자외선차단", "ACTIVE": "기능성활성", "THICKENER": "점증제",
    "COLORANT": "색소", "FRAGRANCE": "향료", "SOLVENT": "용매",
    "PH_ADJUSTER": "pH조절제", "CHELATING": "킬레이트제", "FILM_FORMER": "피막형성제",
    "OTHER": "기타"
}
ORIGIN_KR = {"NATURAL": "천연", "SYNTHETIC": "합성", "BIOTECHNOLOGY": "바이오", "MINERAL": "광물"}

for i, row in enumerate(rows):
    r = START_ROW + 1 + i
    ws1.cell(row=r, column=1, value=i + 1)
    ws1.cell(row=r, column=2, value=row[1])  # inci_name
    ws1.cell(row=r, column=3, value=row[2])  # korean_name
    ws1.cell(row=r, column=4, value=row[3])  # cas_number
    ws1.cell(row=r, column=5, value=TYPE_KR.get(row[4], row[4]))
    ws1.cell(row=r, column=6, value=ORIGIN_KR.get(row[5], row[5] or ""))
    ws1.cell(row=r, column=7, value=row[6])  # description
    ws1.cell(row=r, column=8, value=row[7])  # source
    ws1.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws1.cell(row=r, column=4).alignment = Alignment(horizontal="center", vertical="center")
    ws1.cell(row=r, column=5).alignment = Alignment(horizontal="center", vertical="center")
    ws1.cell(row=r, column=6).alignment = Alignment(horizontal="center", vertical="center")
    ws1.cell(row=r, column=8).alignment = Alignment(horizontal="center", vertical="center")
    style_data(ws1, r, len(headers), alt=(i % 2 == 1))

ws1.auto_filter.ref = f"A{START_ROW}:H{START_ROW + len(rows)}"
ws1.freeze_panes = f"A{START_ROW + 1}"
auto_width(ws1, len(headers))


# ════════════════════════════════════════════════════════════════════════════
# Sheet 2: 규제/안전성 정보
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("규제 및 안전성")
ws2.sheet_properties.tabColor = "C00000"

cur.execute("""
    SELECT kb.search_key, kb.data, im.korean_name, im.ingredient_type
    FROM coching_knowledge_base kb
    LEFT JOIN ingredient_master im ON LOWER(im.inci_name) = kb.search_key
    WHERE kb.category = 'INGREDIENT_REGULATION'
    ORDER BY kb.search_key
""")
kb_rows = cur.fetchall()

add_title(ws2, "원료 규제 및 안전성 정보", f"총 {len(kb_rows)}건 | 추출일: {datetime.now().strftime('%Y-%m-%d %H:%M')}")

headers2 = ["No.", "INCI명", "한국어명", "원료 유형", "최대 사용농도", "EWG 점수",
            "한국 식약처 규제", "EU 규제", "안전성 메모"]
START_ROW2 = 4
for c, h in enumerate(headers2, 1):
    ws2.cell(row=START_ROW2, column=c, value=h)
style_header(ws2, START_ROW2, len(headers2))
ws2.row_dimensions[START_ROW2].height = 30

for i, row in enumerate(kb_rows):
    r = START_ROW2 + 1 + i
    data = row[1] if isinstance(row[1], dict) else json.loads(row[1]) if row[1] else {}

    ws2.cell(row=r, column=1, value=i + 1)
    ws2.cell(row=r, column=2, value=data.get("inci_name", row[0]))
    ws2.cell(row=r, column=3, value=row[2] or "")
    ws2.cell(row=r, column=4, value=TYPE_KR.get(row[3], row[3] or ""))
    ws2.cell(row=r, column=5, value=data.get("max_concentration", ""))
    ewg = data.get("ewg_score")
    ws2.cell(row=r, column=6, value=ewg if ewg else "")
    ws2.cell(row=r, column=7, value=data.get("kr_regulation", ""))
    ws2.cell(row=r, column=8, value=data.get("eu_regulation", ""))
    ws2.cell(row=r, column=9, value=data.get("safety_notes", ""))

    # 센터 정렬
    for cc in [1, 4, 5, 6]:
        ws2.cell(row=r, column=cc).alignment = Alignment(horizontal="center", vertical="center")

    # EWG 점수 색상 (1-2 녹색, 3-6 주황, 7+ 빨강)
    if ewg:
        ewg_cell = ws2.cell(row=r, column=6)
        ewg_cell.alignment = Alignment(horizontal="center", vertical="center")
        if isinstance(ewg, (int, float)):
            if ewg <= 2:
                ewg_cell.font = Font(name="맑은 고딕", size=10, bold=True, color="006100")
                ewg_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif ewg <= 6:
                ewg_cell.font = Font(name="맑은 고딕", size=10, bold=True, color="9C5700")
                ewg_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            else:
                ewg_cell.font = Font(name="맑은 고딕", size=10, bold=True, color="9C0006")
                ewg_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    style_data(ws2, r, len(headers2), alt=(i % 2 == 1))

ws2.auto_filter.ref = f"A{START_ROW2}:I{START_ROW2 + len(kb_rows)}"
ws2.freeze_panes = f"A{START_ROW2 + 1}"
auto_width(ws2, len(headers2))


# ════════════════════════════════════════════════════════════════════════════
# Sheet 3: 규제 캐시 (KR/EU 상세)
# ════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("규제 캐시 상세")
ws3.sheet_properties.tabColor = "ED7D31"

cur.execute("""
    SELECT rc.source, rc.ingredient, rc.inci_name, rc.max_concentration,
           rc.restriction, rc.updated_at
    FROM regulation_cache rc
    ORDER BY rc.source, rc.ingredient
""")
reg_rows = cur.fetchall()

add_title(ws3, "규제 캐시 상세", f"총 {len(reg_rows)}건 | 추출일: {datetime.now().strftime('%Y-%m-%d %H:%M')}")

headers3 = ["No.", "출처", "원료명", "INCI명", "최대 사용농도", "규제/제한 사항", "갱신일"]
START_ROW3 = 4
for c, h in enumerate(headers3, 1):
    ws3.cell(row=START_ROW3, column=c, value=h)
style_header(ws3, START_ROW3, len(headers3))
ws3.row_dimensions[START_ROW3].height = 30

SOURCE_KR = {
    "GEMINI_KR": "🇰🇷 한국 식약처",
    "GEMINI_EU": "🇪🇺 EU CosIng",
    "MFDS_SEED": "🇰🇷 식약처 (시드)",
}

for i, row in enumerate(reg_rows):
    r = START_ROW3 + 1 + i
    ws3.cell(row=r, column=1, value=i + 1)
    ws3.cell(row=r, column=2, value=SOURCE_KR.get(row[0], row[0]))
    ws3.cell(row=r, column=3, value=row[1])
    ws3.cell(row=r, column=4, value=row[2])
    ws3.cell(row=r, column=5, value=row[3])
    ws3.cell(row=r, column=6, value=row[4])
    ws3.cell(row=r, column=7, value=str(row[5])[:19] if row[5] else "")

    for cc in [1, 2, 5, 7]:
        ws3.cell(row=r, column=cc).alignment = Alignment(horizontal="center", vertical="center")
    style_data(ws3, r, len(headers3), alt=(i % 2 == 1))

ws3.auto_filter.ref = f"A{START_ROW3}:G{START_ROW3 + len(reg_rows)}"
ws3.freeze_panes = f"A{START_ROW3 + 1}"
auto_width(ws3, len(headers3))


# ════════════════════════════════════════════════════════════════════════════
# Sheet 4: 수집 현황 요약
# ════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("수집 현황")
ws4.sheet_properties.tabColor = "70AD47"

add_title(ws4, "데이터 수집 현황 요약", f"추출일: {datetime.now().strftime('%Y-%m-%d %H:%M')}")

# 테이블 통계
cur.execute("SELECT count(*) FROM ingredient_master")
cnt_master = cur.fetchone()[0]
cur.execute("SELECT count(*) FROM ingredient_properties")
cnt_props = cur.fetchone()[0]
cur.execute("SELECT count(*) FROM regulation_cache")
cnt_reg = cur.fetchone()[0]
cur.execute("SELECT count(*) FROM coching_knowledge_base WHERE category='INGREDIENT_REGULATION'")
cnt_kb = cur.fetchone()[0]

stats = [
    ["테이블", "건수", "설명", "상태"],
    ["ingredient_master", cnt_master, "원료 기본정보 (INCI명, 한국어명, CAS, 유형)", "● 수집중"],
    ["ingredient_properties", cnt_props, "물리화학적 물성 (분자량, LogP, SMILES 등)", "○ 대기"],
    ["regulation_cache", cnt_reg, "규제 정보 (한국/EU 규제, 사용농도 제한)", "● 수집중"],
    ["coching_knowledge_base", cnt_kb, "규제 상세 JSONB (EWG, 안전성 메모 등)", "● 수집중"],
]

ROW = 4
for c, h in enumerate(stats[0], 1):
    ws4.cell(row=ROW, column=c, value=h)
style_header(ws4, ROW, 4)

for i, s in enumerate(stats[1:]):
    r = ROW + 1 + i
    for c, v in enumerate(s, 1):
        ws4.cell(row=r, column=c, value=v)
    ws4.cell(row=r, column=2).alignment = Alignment(horizontal="center")
    ws4.cell(row=r, column=4).alignment = Alignment(horizontal="center")
    if "● 수집중" in str(s[3]):
        ws4.cell(row=r, column=4).font = Font(name="맑은 고딕", size=10, color="006100", bold=True)
    style_data(ws4, r, 4, alt=(i % 2 == 1))

auto_width(ws4, 4)

# 유형별 분포
cur.execute("SELECT ingredient_type, count(*) FROM ingredient_master GROUP BY ingredient_type ORDER BY count(*) DESC")
type_dist = cur.fetchall()

ROW2 = ROW + len(stats) + 2
ws4.cell(row=ROW2, column=1, value="원료 유형별 분포").font = Font(name="맑은 고딕", bold=True, size=12, color="2F5496")
ROW2 += 1
for c, h in enumerate(["유형 (영문)", "유형 (한글)", "건수"], 1):
    ws4.cell(row=ROW2, column=c, value=h)
style_header(ws4, ROW2, 3)

for i, (tp, cnt) in enumerate(type_dist):
    r = ROW2 + 1 + i
    ws4.cell(row=r, column=1, value=tp)
    ws4.cell(row=r, column=2, value=TYPE_KR.get(tp, tp))
    ws4.cell(row=r, column=3, value=cnt)
    ws4.cell(row=r, column=3).alignment = Alignment(horizontal="center")
    style_data(ws4, r, 3, alt=(i % 2 == 1))


# ── 저장 ─────────────────────────────────────────────────────────────────────
from datetime import datetime as _dt
_stamp = _dt.now().strftime('%Y%m%d_%H%M')
OUTPUT = f"/mnt/e/COCHING/backup/COCHING_원료DB_{_stamp}.xlsx"
wb.save(OUTPUT)
cur.close()
conn.close()
print(f"✅ Excel 저장 완료: {OUTPUT}")
print(f"   원료 마스터: {cnt_master}건")
print(f"   규제/안전성: {cnt_kb}건")
print(f"   규제 캐시: {cnt_reg}건")
